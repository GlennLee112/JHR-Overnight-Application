# utility
import os
import sys
import datetime
from dateutil.relativedelta import relativedelta
import os

# threading / multi processing
from threading import Thread

# tkinter related
import tkinter as tk
from tkinter import (
    StringVar,
    PhotoImage,
    Label,
    OptionMenu,
    Button,
    messagebox, )
from tkinter import ttk
import tkcalendar

# script
import Main_Arrival_Mon
import Main_Delivery_Mon
import Main_Overnight
import Directory.jnt_SQL as jnt_sql
from Directory.File_Manager import file_dir_create

# data / excel
import pandas as pd
from openpyxl import utils, load_workbook


# back button reference:
# https://stackoverflow.com/questions/61681086/how-to-go-back-to-previous-frame-after-button-click-using-tkinter-on-python
# Plan of menu layout

# 0. Paths set up
# Paths


def get_base_path():
    # Ref:
    # https://stackoverflow.com/questions/404744/determining-application-path-in-a-python-exe-generated-by-pyinstaller
    # https://stackoverflow.com/questions/66734987/absolute-path-of-exe-file-not-working-properly-in-python
    """Base path function condition gate to determine if file is running as script of as executable"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    elif __file__:
        return os.path.dirname(__file__)

    # return path_return


# base_path = os.path.dirname(__file__)
base_path = get_base_path()
icon_path = os.path.join(base_path, "tk_res", "jnt_icon.png")
# Input
input_path = os.path.join(base_path, "Input")
input_path_overnight = os.path.join(input_path, "Overnight")
input_path_delivery_mon = os.path.join(input_path, "Delivery Monitoring")
input_path_arrival_mon = os.path.join(input_path, "Arrival Monitoring")
# Output
output_path = os.path.join(base_path, "Output")
output_path_overnight = os.path.join(output_path, "Overnight")
output_path_overnight_summary = os.path.join(output_path, "Overnight", "Summary")
output_path_arrival_mon = os.path.join(output_path, "Arrival Monitoring")
output_path_delivery_mon = os.path.join(output_path, "Delivery Monitoring")
# Report
excel_report_path = os.path.join(base_path, "Excel Report")
excel_report_path_overnight = os.path.join(excel_report_path, "Arrival Overnight Pre-delivery")

# template
xlsx_folder = os.path.join(base_path, "Excel Report")
xlsx_arrival_overnight_path = os.path.join(xlsx_folder, "Arrival Overnight Pre-delivery")
template_file = os.path.join(xlsx_arrival_overnight_path, "柔佛留到派预测件量 template.xlsx")
# delivery mon.
delivery_mon_folder = os.path.join(base_path, "Delivery Mon.")
# database
database_file_path = os.path.join(base_path, "Database", "test.db")


# 0. Main
# Multiple tab for a. Input, b. Output, c. AWB search
class Application(tk.Tk):
    def __init__(self):  # Initialization constructor
        super().__init__()  # initiate on tk.TK; no parameter required
        # Photo path
        icon_path_use = PhotoImage(file=os.path.join(os.path.dirname(__file__), "tk_res", "jnt_icon.png"))
        # window setting
        self.geometry("400x325")
        # Set fixed size for window; for simplicity reason
        self.resizable(False, False)
        self.iconphoto(False, icon_path_use)
        # 1. Configuration
        self.title("Report App - Overnight")
        # self.columnconfigure(0, weight=1)
        # self.columnconfigure(1,  # adjust for widget on second column
        #                      weight=3)
        # weight = proportion of the column to be taken; 3 (out of 4) = 75% of additional space
        # self.rowconfigure(0, weight=1)
        self.notebook = ttk.Notebook(self)
        self.notebook.enable_traversal()

        # 2. Notebook
        # Create tab; this tab will utilize 'notebook' for holding different tab
        # https://www.reddit.com/r/learnpython/comments/hc722k/how_do_i_create_multiple_tabs_in_tkinter/
        # https://www.youtube.com/watch?v=t-lUX5e6fOY&ab_channel=CodersLegacy

        # Add in frame in the form of custom function; custom function are made below
        self.notebook.add(UploadSelect(self), text="Upload")
        self.notebook.add(ReportSelect(self), text="Report")
        self.notebook.add(QuerySelect(self), text="Query")

        # frame1.pack(padx = 5, pady=5)
        # frame2.pack(padx = 5, pady=5)

        self.notebook.pack(expand=True, fill="both")

        self.option_add('*tearOff', False)
        top_menu = MenuTotal(self)
        self.config(menu=top_menu)

    def disable_notebook_tabs(self):
        for i in range(self.notebook.index("end")):
            self.notebook.tab(i, state="disabled")

    def enable_notebook_tabs(self):
        for i in range(self.notebook.index("end")):
            self.notebook.tab(i, state="normal")
        # # Create tab; this tab will utilize 'notebook' for holding different tab
        # # https://www.reddit.com/r/learnpython/comments/hc722k/how_do_i_create_multiple_tabs_in_tkinter/
        # # Notebook for holding tabs
        # notebook = ttk.Notebook(self)
        # Add tab
        # notebook.add(InputForm(self), text="tab 1")
        # notebook.add(Textlist(self), text="tab 2")
        # # Pack tab; mandatory
        # notebook.pack(expand=1, fill='both')


class MenuTotal(tk.Menu):
    """Menu for JHR Overnight App GUI, with the following cascade:
    1. Data
    2. Folder"""
    # https://tkdocs.com/tutorial/menus.html
    # https://www.pythontutorial.net/tkinter/tkinter-menu/
    def __init__(self, root):
        super().__init__(root)
        # 1. Data_Menu
        # Clear DB data
        data_menu = tk.Menu(self)
        data_menu.add_command(label="Clear DB", command=self.clear_db)

        # 2. Folder_menu
        # Folder open option; start the folder selected
        folder_menu = tk.Menu(self)
        # folder_menu.add_command(label="Folder", command=dummy_func)
        # # 2.1 Output submenu
        output_submenu = tk.Menu(self)
        output_submenu.add_command(label="Overnight", command=self.open_out_overnight)
        output_submenu.add_command(label="Delivery Mon.", command=self.open_out_delivery)
        output_submenu.add_command(label="Arrival Mon.", command=self.open_out_arrival)
        # # 2.2 Excel submenu
        excel_submenu = tk.Menu(self)
        excel_submenu.add_command(label="JHR Overnight", command=self.open_report_jhr_overnight)

        # 3. folder submenu cascade
        folder_menu.add_cascade(label="Output", menu=output_submenu)
        folder_menu.add_separator()
        folder_menu.add_cascade(label="Excel Report", menu=excel_submenu)

        # 4. Main cascade setup
        self.add_cascade(menu=data_menu, label="Data")
        self.add_cascade(menu=folder_menu, label="Folder")

    @staticmethod
    def clear_db():
        """standalone function for menu to clear out database at request"""
        msg = (f"Delete AWB DB?\n"
               f"\n"
               f"DB and all AWB data will be deleted")

        answer = tk.messagebox.askyesno(title='Clear DB', message=msg)
        # Logic gate: delete if clicked YES; ignore if clicked NO.
        if answer:
            print("order 66")
            os.remove(database_file_path)

        else:
            print('nothing happened')
            return

    @staticmethod
    def open_out_arrival():
        # overnight_out = os.path.join()
        os.startfile(output_path_arrival_mon)

    @staticmethod
    def open_out_delivery():
        # overnight_out = os.path.join()
        os.startfile(output_path_delivery_mon)

    @staticmethod
    def open_out_overnight():
        # overnight_out = os.path.join()
        os.startfile(output_path_overnight)

    @staticmethod
    def open_report_jhr_overnight():
        # overnight_out = os.path.join()
        os.startfile(excel_report_path_overnight)


class UploadSelect(ttk.Frame):
    def __init__(self, parent):  # Initialization constructor
        # all elements that user wanted to access outside the function must prefix with 'self.'
        super().__init__(parent)

        # 1. Configuration
        # self.columnconfigure(0, weight=1)
        # self.rowconfigure(0, weight=1)
        self.columnconfigure(1, weight=3)
        self.rowconfigure(1, weight=0)

        # 2. Label
        # Label containing text to indicate items in GUI
        # Label reference:
        # https://www.geeksforgeeks.org/python-tkinter-label/
        # 2.1 Assign StringVar for label to use
        # 2.1.1 Main
        upload_info = StringVar()
        upload_info.set("This is the upload menu\n"
                        "\n"
                        "For overnight, please ensure file is in '.csv UTF-8' format")
        # 2.1.2 Step 1 - select
        step_1_info = StringVar()
        step_1_info.set("Step 1: Select (optional: open folder to paste csv file if haven't)")
        # 2.1.3 Step 2 - run
        step_2_info = StringVar()
        step_2_info.set("Step 2: Click 'run'")

        # 2.2 Set up label elements
        # 2.2.1 Label 1
        label_1 = Label(self,
                        textvariable=upload_info,
                        pady=5,
                        padx=5,
                        wraplength=300)

        label_1.grid(row=0, column=1)
        # Change wrap dynamically
        # https://stackoverflow.com/questions/62485520/how-to-wrap-the-text-in-a-tkinter-label-dynamically
        label_1.bind('<Configure>', lambda e: label_1.config(wraplength=(self.winfo_width() - 25)))

        label_2 = Label(self,
                        textvariable=step_1_info,
                        pady=5,
                        padx=5,
                        wraplength=300)

        label_2.grid(row=1, column=1)
        # Change wrap dynamically
        # https://stackoverflow.com/questions/62485520/how-to-wrap-the-text-in-a-tkinter-label-dynamically
        label_2.bind('<Configure>', lambda e: label_2.config(wraplength=(self.winfo_width() - 25)))

        label_3 = Label(self,
                        textvariable=step_2_info,
                        pady=5,
                        padx=5,
                        wraplength=300)

        label_3.grid(row=5, column=1)
        # Change wrap dynamically
        # https://stackoverflow.com/questions/62485520/how-to-wrap-the-text-in-a-tkinter-label-dynamically
        label_3.bind('<Configure>', lambda e: label_2.config(wraplength=(self.winfo_width() - 25)))

        # 3. Selection option
        # 3.1 set up menu and associated function, self.var would be the default variable
        self.var_upload = StringVar(value="All")

        menu = ["All", "General DP Overnight", "Delivery Mon.", "Arrival Mon."]
        OptionMenu(self, self.var_upload, *menu).grid(row=2, column=1, pady=5)
        # 3.2 Button
        # Run button will change tasks based on selection
        self.run_btn = Button(self, text="Run", command=self.run_overnight_selector)
        self.run_btn.grid(row=6, column=1, pady=5)

        # 3.3 File location button
        # Open file location of involved folder base on
        self.folder_btn = Button(self, text="Folder", command=self.folder_open)
        self.folder_btn.grid(row=3, column=1, pady=5)
        self.folder_btn.config(width=8)

    def run_overnight_selector(self):
        # Planned implementation (Completed):
        # 1. dialog pop up
        # 2. selector
        # 3. threading task running and progress tracking using progress bar
        # 4. auto destroy after completion

        # 1. dialog set up
        ongoing_dialog = tk.Toplevel()
        ongoing_dialog.iconphoto(False, PhotoImage(file=icon_path))
        ongoing_dialog.geometry("250x100")

        # 2. Selector
        # Implement threading when running tasks
        # https://stackoverflow.com/questions/49775762/trigger-an-event-when-a-thread-is-finished
        # Set ongoing_dialog to top most
        # https://stackoverflow.com/questions/8691655/how-to-put-a-tkinter-window-on-top-of-the-others
        ongoing_dialog.attributes('-topmost', True)
        selected = self.var_upload.get()

        # 1. Select 'All'
        if selected == "All":
            # call in 'all_run' external function from outside of class
            all_run(ongoing_dialog)

        # 2. Select 'General DP Overnight'
        elif selected == "General DP Overnight":
            # Run General DP  script
            def task():
                try:
                    status_label = Label(ongoing_dialog, text='Ongoing')
                    status_label.pack()
                    run_overnight()
                    ongoing_dialog.after(0, lambda: status_label.config(text='Completed'))
                    ongoing_dialog.after(5000, ongoing_dialog.destroy)
                # Exception handling
                except Exception as e:
                    print("Error during processing:", e)
                    ongoing_dialog.after(0, ongoing_dialog.destroy)
                    # messagebox.showerror("Error", f"An error occurred: {e}")

            Thread(target=task).start()

        # 3. Select 'Delivery Mon.'
        elif selected == "Delivery Mon.":
            # Run Delivery Month script
            def task():
                try:
                    status_label = Label(ongoing_dialog, text='Ongoing')
                    status_label.pack()
                    run_delivery_mon()
                    ongoing_dialog.after(0, lambda: status_label.config(text='Completed'))
                    ongoing_dialog.after(5000, ongoing_dialog.destroy)
                except Exception as e:
                    print("Error during processing:", e)
                    ongoing_dialog.after(0, ongoing_dialog.destroy)
                    # messagebox.showerror("Error", f"An error occurred: {e}")

            Thread(target=task).start()

        # 4. Select 'Arrival Mon.'
        elif selected == "Arrival Mon.":
            # Run arrival Month script
            def task():
                try:
                    status_label = Label(ongoing_dialog, text='Ongoing')
                    status_label.pack()
                    run_arrival_mon()
                    ongoing_dialog.after(0, lambda: status_label.config(text='Completed'))
                    ongoing_dialog.after(5000, ongoing_dialog.destroy)
                except Exception as e:
                    print("Error during processing:", e)
                    ongoing_dialog.after(0, ongoing_dialog.destroy)

            Thread(target=task).start()

    def folder_open(self):
        """Function to open folder base on selection"""
        # 1. Get var_upload variable value from main frame / class.
        selected = self.var_upload.get()
        # 2. Condition gate for which file to start.
        # start file ref:
        # https://www.youtube.com/watch?app=desktop&v=iV5sti2hJJQ&ab_channel=MatadorSoftware
        if selected == "All":
            # List of files to open
            folder_all = [input_path_overnight, input_path_arrival_mon, input_path_delivery_mon]
            # Iterate through list above
            for folder in folder_all:
                os.startfile(folder)

        elif selected == "General DP Overnight":
            os.startfile(input_path_overnight)

        elif selected == "Delivery Mon.":
            os.startfile(input_path_delivery_mon)

        elif selected == "Arrival Mon.":
            os.startfile(input_path_arrival_mon)


def all_run(ongoing_dialog):
    status_label = Label(ongoing_dialog, text='Ongoing')
    status_label.pack()

    def task():
        try:
            run_overnight()
            run_delivery_mon()
            run_arrival_mon()
            ongoing_dialog.after(0, lambda: status_label.config(text='Completed'))
            ongoing_dialog.after(5000, ongoing_dialog.destroy)
        except Exception as e:
            print("Error during processing:", e)
            ongoing_dialog.after(0, ongoing_dialog.destroy)
            messagebox.showerror("Error", f"An error occurred: {e}")

    Thread(target=task).start()
    # task()


def run_overnight():
    """independent function that runs the overnight.main script

        To be called in after clicking 'run' button with "General DP Overnight" selected
        """
    # #Method - import py Script and run function + threading
    # overnight_script = Main_Overnight.main
    # overnight_thread = Thread(target=overnight_script)  # assign thread
    # overnight_thread.start()  # start thread

    # Moved thread implementation to 'run_overnight_selector'
    try:
        Main_Overnight.main()
    except Exception as e:
        print("Error during processing:", e)
        messagebox.showerror("Error", f"An error occurred: {e}")


def run_delivery_mon():
    """independent function that runs the overnight.main script

        To be called in after clicking 'run' button with "Delivery Mon." selected
        """
    # #Method - import py Script and run function + threading
    # overnight_script = Main_Overnight.main
    # overnight_thread = Thread(target=overnight_script)  # assign thread
    # overnight_thread.start()  # start thread

    # Moved thread implementation to 'run_overnight_selector'
    try:
        Main_Delivery_Mon.main()
    except Exception as e:
        print("Error during processing:", e)
        messagebox.showerror("Error", f"An error occurred: {e}")


def run_arrival_mon():
    """independent function that runs the arrival.main script

        To be called in after clicking 'run' button
        """
    try:
        Main_Arrival_Mon.main()
    except Exception as e:
        print("Error during processing:", e)
        messagebox.showerror("Error", f"An error occurred: {e}")


class ReportSelect(ttk.Frame):
    def __init__(self, parent):  # Initialization constructor
        # all elements that user wanted to access outside the function must prefix with 'self.'
        super().__init__(parent)
        # 1. Main label
        label_main_text = StringVar()
        label_main_text_var = ("Report generate interface, please select date of the report and then click run to "
                               "generate JHR Overnight Report\n"
                               "\n"
                               "(I.e.: on 2nd May 2025, select 5/2/2025)")
        label_main_text.set(label_main_text_var)
        self.label_main = Label(self, textvariable=label_main_text)
        self.label_main.grid(row=0, column=0)
        self.label_main.bind('<Configure>', lambda e: self.label_main.config(wraplength=(self.winfo_width() - 25)))

        # 2. Instances
        # self.error_log = []

        # 2.1 Create an instances of inner_frame; repeat for all the frame required
        self.inner_frame_1 = self.InnerFrameDate(self)
        self.inner_frame_1.grid(row=2, column=0)

        # 2.2 Instance for selected date, to be used for function to obtain report date required
        # self.selected_date = ""

        # 3. Widget
        # 3.1 Folder Button
        # self.button_run = Button(self, text="Folder Report", command=self.print_main_date)
        # self.button_run.grid(row=3, column=0)
        # Run button
        # self.button_run = Button(self, text="Generate Report")
        # self.button_run.grid(row=2, column=1)

        # self.pack()

    # def print_main_date(self):
    #     self.selected_date = self.inner_frame_1.cld_1.get_date()
    #     print(self.selected_date)

    class InnerFrameDate(ttk.Frame):
        def __init__(self, parent):
            super().__init__(parent)
            # 1. label
            # 1.1 Main label
            # label_main_text = StringVar()
            # label_main_text_var = "Step 1:"
            # label_main_text.set(label_main_text_var)
            # self.label_main = Label(self, textvariable=label_main_text)
            # self.label_main.grid(row=0, column=0)

            # 1.2 Select date label
            label_1_text = StringVar()
            label_1_text.set("Step 1: Select Date")
            self.label_1 = Label(self, textvariable=label_1_text, justify="left", padx=3, pady=3)
            self.label_1.grid(row=1, column=0)

            # 1.3 Run or edit template label
            label_2_text = StringVar()
            label_2_text.set("Step 2: Run or edit template")
            self.label_2 = Label(self, textvariable=label_2_text, justify="left", padx=3, pady=3)
            self.label_2.grid(row=2, column=0)

            # 1.4 Run or edit label
            label_2_text = StringVar()
            label_2_text.set("Step 3: Open generated file in folder")
            self.label_2 = Label(self, textvariable=label_2_text, justify="left", padx=3, pady=3)
            self.label_2.grid(row=4, column=0)

            # 2. widget
            # 2.1 date entry widget
            self.cld_1 = tkcalendar.DateEntry(self, locale="en_US", maxdate=datetime.datetime.today().date())
            self.cld_1.grid(row=1, column=1)
            # 2.2 Run button
            self.button_run = Button(self, text="Generate Report", command=self.generate_report, padx=3, pady=3)
            self.button_run.grid(row=2, column=1)
            # 2.3 Edit template button
            self.button_template = Button(self, text="Template Edit", command=self.open_template_file, padx=3, pady=3)
            self.button_template.grid(row=3, column=1)
            # 2.4 date selector
            self.new_file = Button(self, text="New File", command=self.new_file_open, padx=3, pady=3)
            self.new_file.grid(row=4, column=1)

            # 3. Log / variable hold
            self.selected_date = StringVar(value="")
            self.new_file_loc = StringVar(value="")
            # self.error_log = None

            # self.cld_2 = tkcalendar.DateEntry(self, locale="en_US")
            # self.cld_2.grid(row=1, column=1)

        def print_select_date(self):
            print(self.selected_date)

        def print_date(self):
            self.selected_date = self.cld_1.get_date()
            print(self.selected_date)
            # return date

        @staticmethod
        def open_template_file():
            os.startfile(template_file)

        def generate_report(self):
            ongoing_dialog = tk.Toplevel()
            ongoing_dialog.iconphoto(False, PhotoImage(file=icon_path))
            ongoing_dialog.geometry("250x100")
            # Set ongoing_dialog to top most
            # https://stackoverflow.com/questions/8691655/how-to-put-a-tkinter-window-on-top-of-the-others
            ongoing_dialog.attributes('-topmost', True)

            def report_task():
                status_label = Label(ongoing_dialog, text='Ongoing')
                status_label.pack()
                """Generate JHR Overnight Delivery Arrival Pre-delivery report"""
                # 0. Paths
                # 1. Get OG date
                # Used by report title and querying for today overnight
                date = self.cld_1.get_date()
                print(f'selected date is {date}')
                print(f'selected date date type is {type(date)}')

                # 2. Date processing
                # 2.1 Today date
                # Used by Today overnight and report out path
                report_date_year = date.strftime("%Y")
                report_date_year_month = date.strftime("%Y%m")
                report_date = date.strftime("%Y-%m-%d")
                report_date_title = date.strftime("%Y.%m.%d")

                # 2.2 Yesterday
                # Used by Arrival, yesterday overnight, and Delivery Monitoring query
                report_y_date = date - relativedelta(days=1)
                print(f'Report date for yesterday overnight is {report_y_date}')
                report_y_date_year = report_y_date.strftime("%Y")
                report_y_date_year_month = report_y_date.strftime("%Y%m")
                report_y_date_full = report_y_date.strftime('%Y-%m-%d')
                report_y_date_title = report_y_date.strftime("%Y.%m.%d")

                # Error log variable
                # error_log = ""

                # 3. Query
                try:
                    # 3.1 Arrival mon. query (csv)
                    print("Arrival Querying")
                    arr_mon_query_path = os.path.join(output_path_arrival_mon, report_y_date_year,
                                                      report_y_date_year_month + " Arrival Mon.csv")
                    # print(arr_mon_query_path) # for debugging; enable if required
                    # print(os.path.isfile(arr_mon_query_path)) # for debugging; enable if required
                    arr_mon_query_df = pd.read_csv(str(arr_mon_query_path), index_col=False)
                    # filter by date
                    arr_mon_query_df_filtered = (arr_mon_query_df[arr_mon_query_df['Data Date'] == report_y_date_full]
                                                 .reset_index())
                    arr_mon_query_df_filtered = arr_mon_query_df_filtered[['Operating Station No.',
                                                                           '应到总票数',
                                                                           'Qty | Arrived',
                                                                           'Data Date']]
                    # print(arr_mon_query_df_filtered)
                    # print(arr_mon_query_df_filtered)
                    # print(arr_mon_query_df_filtered.empty)
                    # error_log = "Arrival"

                    # if arr_mon_query_df_filtered.empty:
                    #     print(f"Debug - error_log set to: {self.error_log}")
                    #     raise Exception("Empty Data Frame")

                    # 3.2 Delivery mon. query (csv)
                    print("Delivery Querying")
                    del_mon_query_path = os.path.join(output_path_delivery_mon, report_y_date_year,
                                                      report_y_date_year_month + ".csv")
                    # print(del_mon_query_path) # for debugging; enable if required
                    # print(os.path.isfile(del_mon_query_path)) # for debugging; enable if required
                    del_mon_query_df = pd.read_csv(str(del_mon_query_path), index_col=False)
                    # print(del_mon_query_df)
                    # filter by date
                    del_mon_query_df_filtered = (del_mon_query_df[del_mon_query_df['Data Date'] == report_y_date_full]
                                                 .reset_index())

                    del_mon_query_df_filtered = del_mon_query_df_filtered[['DP No. | Delivery',
                                                                           'Dispatcher ID',
                                                                           'Volume | Delivery',
                                                                           'Volume | Delivery Signature',
                                                                           'Data Date']]
                    # print(del_mon_query_df_filtered.empty)
                    # print(del_mon_query_df_filtered)

                    # 3.3 Overnight Summary query (csv)
                    overnight_overall_query_path_y = os.path.join(output_path_overnight_summary, report_y_date_year,
                                                                  report_y_date_year_month +
                                                                  ".csv")
                    print(overnight_overall_query_path_y)
                    print(os.path.isfile(overnight_overall_query_path_y))
                    # ## 3.3.1 Overnight Yesterday
                    print("Overnight Yesterday Querying")
                    overnight_over_query_df_y = pd.read_csv(str(overnight_overall_query_path_y), index_col=False)
                    # filter by date
                    overnight_over_query_df_y_filtered = (overnight_over_query_df_y[overnight_over_query_df_y
                                                                                    ['Data Date']
                                                                                    == report_y_date_full]
                                                          .reset_index()  # reset index
                                                          .sort_values(by=['Scanning DP No. | Last']))  # sort value

                    overnight_over_query_df_y_filtered = overnight_over_query_df_y_filtered[['Scanning DP No. | Last',
                                                                                             'AWB']]
                    print(overnight_over_query_df_y_filtered.empty)
                    print(overnight_over_query_df_y_filtered)

                    # ## 3.3.2 Overnight Today
                    print("Overnight Today Querying")
                    overnight_overall_query_path = os.path.join(output_path_overnight_summary, report_date_year,
                                                                report_date_year_month + ".csv")

                    overnight_over_query_df = pd.read_csv(str(overnight_overall_query_path), index_col=False)
                    # filter by date
                    overnight_over_query_df_filtered = (overnight_over_query_df[overnight_over_query_df['Data Date']
                                                                               == report_date]
                                                        .reset_index()  # reset index; else 'iterrow' will not function
                                                        .sort_values(by=['Scanning DP No. | Last']))  # Sort value

                    overnight_over_query_df_filtered = overnight_over_query_df_filtered[['Scanning DP No. | Last',
                                                                                         'AWB']]
                    print(overnight_over_query_df_filtered.empty)
                    print(overnight_over_query_df_filtered)

                    # 3.4 Overnight AWB query
                    print("Overnight AWB Querying")
                    overnight_over_query_awb = jnt_sql.overnight_awb_read(report_date, database_file_path)
                    print(overnight_over_query_awb.empty)
                    print(overnight_over_query_awb)

                # exception handling, display error
                except Exception as e:
                    ongoing_dialog.after(0, lambda: status_label.config(text=
                                                                        f'Datas for {date} contains error\n'
                                                                        f'\n'
                                                                        f'Please update data or try another date'))

                    ongoing_dialog.after(5000, ongoing_dialog.destroy)

                    return

                # 4. Report writing
                # 4.1 Copy and rename template file to correct folder
                jhr_overnight_out_dir = os.path.join(xlsx_arrival_overnight_path, report_date_year,
                                                     report_date_year_month)
                jhr_overnight_out_path = os.path.join(xlsx_arrival_overnight_path, report_y_date_year,
                                                      report_y_date_year_month, report_y_date_title +
                                                      ' 柔佛留到派预测件量 Johor Overnight Arrival Pre-Delivery summary.xlsx')

                # Load template file and copy formatted file to correct folder
                try:
                    template_workbook = load_workbook(template_file)
                    print('template read complete')
                    # str path from jhr_overnight_out_path variable
                    file_dir_create(jhr_overnight_out_dir)
                    print('Directory create complete')
                    template_workbook.save(str(jhr_overnight_out_path))
                    print('Template file save completed')
                except PermissionError:
                    ongoing_dialog.after(0, lambda: status_label.config(text=
                                                                        f'Template file is still opened\n'
                                                                        f'Please close the template file'))

                    ongoing_dialog.after(5000, ongoing_dialog.destroy)

                    return

                # 4.2 Write
                # 4.2.0 preparation
                # temp solution: ExcelWriter
                # Future solution: for efficiency purpose, use openpyxl directly
                # https://stackoverflow.com/questions/66645235/how-to-write-multiple-pandas-dataframes-to-excel-current-method-corrupts-xlsx
                # Convert singular report_date value to dataframe
                report_date_df = pd.DataFrame({'col1': [report_date]})

                # 4.2.1 df and sheet / col / row value
                # df list
                df_list = [report_date_df,
                           arr_mon_query_df_filtered,
                           del_mon_query_df_filtered,
                           overnight_over_query_df_y_filtered,
                           overnight_over_query_df_filtered,
                           overnight_over_query_awb]

                # sheet / col / row value
                sheet_list_no = [0, 5, 4, 3, 3, 3]
                col_start = [2, 1, 1, 14, 17, 1]
                row_start = [0, 1, 1, 2, 2, 2]
                # wb report out path
                wb_report = load_workbook(jhr_overnight_out_path)

                # 4.2.2 Writing main cluster
                for df, sheet, col_use, row_use in zip(df_list, sheet_list_no, col_start, row_start):
                    ws = wb_report.worksheets[sheet]
                    # Iterate through df rows and populate sheets
                    for idx, row in df.iterrows():
                        for col in range(len(row)):
                            column_letter = utils.get_column_letter(col_use + col)
                            try:
                                ws[f'{column_letter}{idx + row_use + 1}'] = df.iloc[idx, col]
                            except ValueError:
                                continue

                # 4.2.3 Save file as required by openpyxl
                try:
                    wb_report.save(jhr_overnight_out_path)
                except PermissionError:
                    ongoing_dialog.after(0, lambda: status_label.config(text=
                                                                        f'Final file already exists and is opened\n'
                                                                        f'\n'
                                                                        f'Please retry or skip'))

                    ongoing_dialog.after(5000, ongoing_dialog.destroy)

                    return

                print('Delivery: Excel write complete')
                # # Assign new variable to new_file_loc variable; for loading path
                self.new_file_loc = StringVar(value=jhr_overnight_out_path)
                print('File loc write complete')
                # # notification pop up menu
                ongoing_dialog.after(0, lambda: status_label.config(text=
                                                                    f'File for {date} generate successfully'))
                ongoing_dialog.after(5000, ongoing_dialog.destroy)
                # print(date)

            # Start thread
            Thread(target=report_task).start()

        def new_file_open(self):
            new_file_loc_get = self.new_file_loc.get()

            if new_file_loc_get == '':
                # print('no new file') # For debugging
                tk.messagebox.showinfo(title="No file", message="No file generated, please try generate a file first")

            else:
                # print(new_file_loc_get) # For debugging
                new_file_loc_folder = os.path.abspath(new_file_loc_get)
                # print(new_file_loc_folder)
                # print('new file get')
                os.startfile(new_file_loc_folder)

    class InnerFrameGenerate(ttk.Frame):
        def __init__(self, parent):
            super().__init__(parent)

            # 1. Select date label
            label_1_text = StringVar()
            label_1_text.set("Select Date:")
            self.label_1 = Label(self, textvariable=label_1_text)
            self.label_1.grid(row=1, column=0)

            self.cld_1 = tkcalendar.DateEntry(self, locale="en_US")
            self.cld_1.grid(row=1, column=1)

            # label_2_text = StringVar()
            # label_2_text.set("End:")
            # self.label_2 = Label(self, textvariable=label_2_text)
            # self.label_2.grid(row=1, column=0)

            self.button_run = Button(self, text="Generate Report", command=self.print_date)
            self.button_run.grid(row=2, column=0)

            # date selector
            self.selected_date = self.cld_1.get_date()

            # self.cld_2 = tkcalendar.DateEntry(self, locale="en_US")
            # self.cld_2.grid(row=1, column=1)

        def print_date(self):
            date = self.cld_1.get_date()
            print(date)
            return date

        def generate_report(self):
            """Generate JHR Overnight Delivery Arrival Pre-delivery report"""
            date = self.cld_1.get_date()
            print(date)


class QuerySelect(ttk.Frame):
    def __init__(self, parent):  # Initialization constructor
        # all elements that user wanted to access outside the function must prefix with 'self.'
        super().__init__(parent)
        # self.error_log = []
        self.inner_frame = self.InnerFrameDate(self)
        self.inner_frame.grid(row=0, column=1)

        # Pack
        self.pack()

    class InnerFrameDate(ttk.Frame):
        def __init__(self, parent):
            super().__init__(parent)
            # label
            label_1_text = StringVar()
            label_1_text.set("Start:")
            self.label_1 = Label(self, textvariable=label_1_text)
            self.label_1.grid(row=0, column=0)

            label_2_text = StringVar()
            label_2_text.set("End:")
            self.label_2 = Label(self, textvariable=label_2_text)
            self.label_2.grid(row=1, column=0)

            # date selector
            self.cld_1 = tkcalendar.DateEntry(self, locale="en_US")
            self.cld_1.grid(row=0, column=1)

            self.cld_2 = tkcalendar.DateEntry(self, locale="en_US")
            self.cld_2.grid(row=1, column=1)


if __name__ == "__main__":
    app = Application()
    app.mainloop()
